from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from RPA_APP.config import SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASWORD
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import smtplib
import openpyxl
import os
import re
import sys
import json
import pytz


# >>>>>>>>>Money patterns for regex>>>>>>>>>
money_patterns = [
    r"\$\d{1,3}(,\d{3})*(\.\d{1,2})?",   # Format $11.1 ou $111,111.11
    r"\d+ dollars",                      # Format 11 dollars
    r"\d+ USD"                           # Format 11 USD
]
# <<<<<<<<<Money patterns for regex<<<<<<<<<


class Screens():

    # >>>>>>>>>Combine data creation and excel file generation into one function>>>>>>>>>
    def make_excel_file(self, payload, data, directory, excel_filename):
        try:
            # >>>>>>>>>Prepare the data for the excel file>>>>>>>>>
            count = 0
            excel_data = []
            for h in data:
                try:
                    count += 1
                    aljazeera_data = [
                        h[0],  # Title
                        datetime.strptime(h[3].replace("Last update ", ""), "%d %b %Y") if len(h) > 3 else None,  # Date
                        h[1],  # Description
                        f"{count}-Aljazeera",  # Picture_Filename
                        h[4] if len(h) > 3 else h[2],  # Picture_URL
                        payload["search_phrase"], # Search_Phrase
                        h[0].upper().count(payload["search_phrase"].upper()) + h[1].upper().count(payload["search_phrase"].upper()),  # Count of "search_phrase" in the title and description
                        True if any(re.search(pattern, h[0]) for pattern in money_patterns) or any(re.search(pattern, h[1]) for pattern in money_patterns) else False,  # Money
                        datetime.now(pytz.timezone("America/Sao_Paulo")).strftime("%d %B %Y %H:%M:%S"),  # DT_Insert
                        datetime.now(pytz.timezone("America/Sao_Paulo")).strftime("%d %B %Y %H:%M:%S")  # DT_Update
                    ]
                    excel_data.append(aljazeera_data)
                    continue
                except Exception as e:
                    print(f"Error processing item {h}: {e}")
                    continue
            # <<<<<<<<<Prepare the data for the excel file<<<<<<<<<

            # >>>>>>>>>Create a new workbook and select the first tab renaming it>>>>>>>>>
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Aljazeera-News"
            # <<<<<<<<<Create a new workbook and select the first tab renaming it<<<<<<<<<

            # >>>>>>>>>Set the header>>>>>>>>>
            header = [
                "Title", "Date", "Description", "Picture_Filename",
                "Picture_URL", "Search_Phrase", "Count_Search_Phrase", "Money",
                "DT_Insert", "DT_Update"
            ]
            ws.append(header)
            # Freeze header
            ws.freeze_panes = "B2"
            # <<<<<<<<<Set the header<<<<<<<<<

            # >>>>>>>>>Format header (bold, black background, white font)>>>>>>>>>
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")  # White font color
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black background
            # <<<<<<<<<Format header (bold, black background, white font)<<<<<<<<<

            # >>>>>>>>>Insert data from the second line>>>>>>>>>
            for row in excel_data:
                ws.append(row)
            # <<<<<<<<<Insert data from the second line<<<<<<<<<

            # >>>>>>>>>Center align and add borders to all cells>>>>>>>>>
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            # <<<<<<<<<Center align and add borders to all cells<<<<<<<<<

            # >>>>>>>>>Auto-adjust column widths based on content>>>>>>>>>
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Get the column letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Add a small buffer
                ws.column_dimensions[col_letter].width = adjusted_width
            # <<<<<<<<<Auto-adjust column widths based on content<<<<<<<<<

            # >>>>>>>>>Check if directory exists, otherwise create>>>>>>>>>
            os.makedirs(directory, exist_ok=True)
            # <<<<<<<<<Check if directory exists, otherwise create<<<<<<<<<

            # >>>>>>>>>Save the file to the specified path>>>>>>>>>
            file_path = os.path.join(directory, excel_filename)
            wb.save(file_path)
            # <<<<<<<<<Save the file to the specified path<<<<<<<<<

            return {"status": True}
        except Exception as e:
            # >>>>>>>>>Tracing the error>>>>>>>>>
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback_details = {
                "filename": exc_traceback.tb_frame.f_code.co_filename,
                "line_number": exc_traceback.tb_lineno,
                "function_name": exc_traceback.tb_frame.f_code.co_name,
                "exception_type": exc_type.__name__,
                "exception_message": str(exc_value)
            }
            print("=-==-==-=ERROR=-==-==-=")
            print(traceback_details)
            print("=-==-==-=ERROR=-==-==-=")
            # <<<<<<<<<Tracing the error<<<<<<<<<
            return {"status": False, "msg": json.dumps(traceback_details)}
    # <<<<<<<<<Combine data creation and excel file generation into one function<<<<<<<<<

    # >>>>>>>>>Send aljazeera excel by email>>>>>>>>>
    def send_excel_email(self, email, subject, directory, excel_filename):
        try:
            # >>>>>>>>>SMTP server settings>>>>>>>>>
            smtp_host = SMTP_HOST
            smtp_port = SMTP_PORT
            smtp_user = SMTP_USER
            smtp_password = SMTP_PASWORD
            # <<<<<<<<<SMTP server settings<<<<<<<<<

            # >>>>>>>>>Create email message>>>>>>>>>
            msg = MIMEMultipart()
            msg['From'] = smtp_user
            msg['To'] = email
            msg['Subject'] = subject
            # <<<<<<<<<Create email message<<<<<<<<<

            # >>>>>>>>>Email body>>>>>>>>>
            body = 'Hello,\n\nAttached is the requested Excel file with the results of the most recent news on the requested subject.'
            msg.attach(MIMEText(body, 'plain'))
            # <<<<<<<<<Email body<<<<<<<<<

            # >>>>>>>>>Attach Excel file>>>>>>>>>
            with open(os.path.join(directory, excel_filename), 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(os.path.join(directory, excel_filename))}')
                msg.attach(part)
            # <<<<<<<<<Attach Excel file<<<<<<<<<

            # >>>>>>>>>Send email>>>>>>>>>
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_password)
                server.sendmail(smtp_user, email, msg.as_string())
            # <<<<<<<<<Send email<<<<<<<<<
            return {"status": True}
        except Exception as e:
            # >>>>>>>>>Tracing the Error>>>>>>>>>
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback_details = {
                "filename": exc_traceback.tb_frame.f_code.co_filename,
                "line_number": exc_traceback.tb_lineno,
                "function_name": exc_traceback.tb_frame.f_code.co_name,
                "exception_type": exc_type.__name__,
                "exception_message": str(exc_value)
            }
            print("=-==-==-=ERROR=-==-==-=")
            print(traceback_details)
            print("=-==-==-=ERROR=-==-==-=")
            # <<<<<<<<<Tracing the Error<<<<<<<<<
            return {"status": False, "msg": json.dumps(traceback_details)}
    # <<<<<<<<<Send aljazeera excel by email<<<<<<<<<
